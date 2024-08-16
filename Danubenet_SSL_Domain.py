import pandas as pd
import ssl
import socket
from datetime import datetime, timezone
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Function to fetch SSL certificate expiration
def get_ssl_expiration(domain):
    try:
        conn = ssl.create_default_context().wrap_socket(socket.socket(), server_hostname=domain)
        conn.settimeout(10.0)
        conn.connect((domain, 443))
        cert = conn.getpeercert()
        conn.close()
        expiration_date_str = cert['notAfter']
        expiration_date = datetime.strptime(expiration_date_str, '%b %d %H:%M:%S %Y GMT')
        expiration_date = expiration_date.replace(tzinfo=timezone.utc)
        days_until_expiration = (expiration_date - datetime.now(timezone.utc)).days
        return days_until_expiration, expiration_date.strftime('%Y-%m-%d')
    except Exception as e:
        return 'Error', str(e)

# Function to normalize domain names
def normalize_domain(domain):
    if isinstance(domain, str):
        return re.sub(r'^\*\.|^\.', '', domain)
    return domain

# Function to update SSL expiry based on the SSL certificate
def update_ssl_expiry(row):
    ssl_cert = row['SSL certificate']
    base_domain = normalize_domain(ssl_cert)
    if isinstance(base_domain, str) and base_domain:
        days_until_expiration, expiration_date = get_ssl_expiration(base_domain)
        if days_until_expiration == 'Error':
            print(f'Error fetching SSL expiration for domain: {ssl_cert}')
            return 'Error'
        print(f'SSL Domain: {ssl_cert}, SSL Expiry Date: {expiration_date}')
        return expiration_date
    return 'Invalid Domain'

# Function to determine status
def determine_status(expiry_date):
    today = datetime.now(timezone.utc)
    if expiry_date and expiry_date != 'Error':
        try:
            expiry_date = datetime.strptime(expiry_date, '%Y-%m-%d')
            expiry_date = expiry_date.replace(tzinfo=timezone.utc)
            days_left = (expiry_date - today).days
            if days_left < 0:
                return 'Expired'
            elif days_left <= 30:
                return 'Near Expiry'
            else:
                return f'{days_left} days left'
        except ValueError:
            return 'Invalid Date'
    return 'No Date'

# Define color fills for Expiry Dates and Status
fill_expired = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
fill_near_expiry = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
fill_valid = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
fill_no_date = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White

# Path to your Excel file
excel_file = r'C:\Users\gauri.kothekar\OneDrive - Netsmartz LLC\Danubenet_SSL_Domain.xlsx'

# Read the Excel file
wb = load_workbook(excel_file)

# Update SSL expiry dates and status in the 'ssl' sheet
if 'ssl' in wb.sheetnames:
    ws_ssl = wb['ssl']
    
    # Convert the sheet to a DataFrame
    data = ws_ssl.values
    columns = next(data)[0:]  # Get the first row for the column names
    df_ssl = pd.DataFrame(data, columns=columns)

    # Ensure required columns exist
    if 'SSL certificate' not in df_ssl.columns:
        raise ValueError("The 'SSL certificate' column is missing from the SSL sheet.")
    if 'SSL Expiry Date' not in df_ssl.columns:
        df_ssl['SSL Expiry Date'] = ''

    # Update the SSL expiry dates
    df_ssl['SSL Expiry Date'] = df_ssl.apply(update_ssl_expiry, axis=1)

    # Update the Excel sheet with the new SSL expiry dates
    for idx, row in df_ssl.iterrows():
        ws_ssl.cell(row=idx + 2, column=df_ssl.columns.get_loc('SSL Expiry Date') + 1, value=row['SSL Expiry Date'])

    # Update Status column in the 'ssl' sheet
    if 'Status' not in df_ssl.columns:
        df_ssl['Status'] = ''
    df_ssl['Status'] = df_ssl['SSL Expiry Date'].apply(determine_status)
    for idx, row in df_ssl.iterrows():
        ws_ssl.cell(row=idx + 2, column=df_ssl.columns.get_loc('Status') + 1, value=row['Status'])
    
    # Apply color fills based on status for SSL Expiry Date and Status
    for idx, row in df_ssl.iterrows():
        status = row['Status']
        expiry_date_cell = ws_ssl.cell(row=idx + 2, column=df_ssl.columns.get_loc('SSL Expiry Date') + 1)
        status_cell = ws_ssl.cell(row=idx + 2, column=df_ssl.columns.get_loc('Status') + 1)
        
        if status == 'Expired':
            expiry_date_cell.fill = fill_expired
            status_cell.fill = fill_expired
        elif status == 'Near Expiry':
            expiry_date_cell.fill = fill_near_expiry
            status_cell.fill = fill_near_expiry
        elif 'days left' in status:
            expiry_date_cell.fill = fill_valid
            status_cell.fill = fill_valid
        else:
            expiry_date_cell.fill = fill_no_date
            status_cell.fill = fill_no_date

# Update Status column in the 'domain' sheet
if 'domain' in wb.sheetnames:
    ws_domain = wb['domain']
    
    # Convert the sheet to a DataFrame
    data = ws_domain.values
    columns = next(data)[0:]  # Get the first row for the column names
    df_domain = pd.DataFrame(data, columns=columns)

    # Ensure required columns exist
    if 'Expiry Date' not in df_domain.columns:
        raise ValueError("The 'Expiry Date' column is missing from the domain sheet.")
    if 'Status' not in df_domain.columns:
        df_domain['Status'] = ''

    # Update the Status column based on Expiry Date
    df_domain['Status'] = df_domain['Expiry Date'].apply(determine_status)
    for idx, row in df_domain.iterrows():
        ws_domain.cell(row=idx + 2, column=df_domain.columns.get_loc('Status') + 1, value=row['Status'])

    # Apply color fills based on status for Expiry Date and Status
    for idx, row in df_domain.iterrows():
        status = row['Status']
        expiry_date_cell = ws_domain.cell(row=idx + 2, column=df_domain.columns.get_loc('Expiry Date') + 1)
        status_cell = ws_domain.cell(row=idx + 2, column=df_domain.columns.get_loc('Status') + 1)
        
        if status == 'Expired':
            expiry_date_cell.fill = fill_expired
            status_cell.fill = fill_expired
        elif status == 'Near Expiry':
            expiry_date_cell.fill = fill_near_expiry
            status_cell.fill = fill_near_expiry
        elif 'days left' in status:
            expiry_date_cell.fill = fill_valid
            status_cell.fill = fill_valid
        else:
            expiry_date_cell.fill = fill_no_date
            status_cell.fill = fill_no_date

# Save the workbook
wb.save(excel_file)
print('Excel file updated with matching color formatting for Expiry Dates and Status successfully.')
