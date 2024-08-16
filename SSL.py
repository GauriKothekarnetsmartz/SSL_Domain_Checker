import pandas as pd
import ssl
import socket
from datetime import datetime, timedelta, timezone
import re
import whois
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import time
import requests
from bs4 import BeautifulSoup

# Function to fetch SSL certificate expiration
def get_ssl_expiration(domain):
    try:
        conn = ssl.create_default_context().wrap_socket(socket.socket(), server_hostname=domain)
        conn.settimeout(10.0)
        conn.connect((domain, 443))
        cert = conn.getpeercert()
        conn.close()
        expiration_date_str = cert['notAfter']
        expiration_date = datetime.strptime(expiration_date_str, '%b %d %H:%M:%S %Y GMT')
        expiration_date = expiration_date.replace(tzinfo=timezone.utc)
        days_until_expiration = (expiration_date - datetime.now(timezone.utc)).days
        return days_until_expiration, expiration_date.strftime('%Y-%m-%d')
    except Exception as e:
        return 'Error', str(e)

# Function to fetch domain registration expiration with retry logic
def get_domain_expiration(domain, retries=3, delay=5):
    for attempt in range(retries):
        try:
            w = whois.whois(domain)
            expiration_date = w.expiration_date
            
            if expiration_date is None:
                raise ValueError("Expiration date not found")
            
            if isinstance(expiration_date, list):
                expiration_date = expiration_date[0]
            
            expiration_date = expiration_date.replace(tzinfo=timezone.utc)
            days_until_expiration = (expiration_date - datetime.now(timezone.utc)).days
            return days_until_expiration, expiration_date.strftime('%Y-%m-%d')
        except Exception as e:
            print(f"Attempt {attempt + 1} failed for {domain}: {e}")
            time.sleep(delay)
    
    # Fallback: Check domain expiration on What's My DNS
    return check_expiration_whatsmydns(domain)

# Function to scrape expiration date from What's My DNS
def check_expiration_whatsmydns(domain):
    try:
        url = f"https://www.whatsmydns.net/domain-expiration?q={domain}"
        response = requests.get(url)
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Parse the expiration date
        date_div = soup.find("div", class_="expiration-date")
        if date_div:
            expiration_date_str = date_div.text.strip()
            expiration_date = datetime.strptime(expiration_date_str, '%Y-%m-%d')
            expiration_date = expiration_date.replace(tzinfo=timezone.utc)
            days_until_expiration = (expiration_date - datetime.now(timezone.utc)).days
            return days_until_expiration, expiration_date.strftime('%Y-%m-%d')
        else:
            raise ValueError("Could not find expiration date on What's My DNS")
    except Exception as e:
        return 'Error', str(e)

# Function to normalize domain names
def normalize_domain(domain):
    if isinstance(domain, str):
        return re.sub(r'^\*\.|^\.', '', domain)
    return domain

# Function to update expiry based on the product type and print information
def update_expiry(row):
    domain = row['Domain']
    prod_type = row['Prod Type']
    base_domain = normalize_domain(domain)
    if prod_type == 'SSL':
        if isinstance(base_domain, str) and base_domain:
            days_until_expiration, expiration_date = get_ssl_expiration(base_domain)
            if days_until_expiration == 'Error':
                print(f'Error fetching SSL expiration for domain: {domain}')
                return 'Error'
            print(f'Domain: {domain}, Expiry Date: {expiration_date}')
            return expiration_date
    elif prod_type == 'Domain':
        if isinstance(base_domain, str) and base_domain:
            days_until_expiration, expiration_date = get_domain_expiration(base_domain)
            if days_until_expiration == 'Error':
                print(f'Error fetching domain expiration for domain: {domain}. Trying What\'s My DNS...')
                days_until_expiration, expiration_date = check_expiration_whatsmydns(base_domain)
                if days_until_expiration == 'Error':
                    print(f'Error fetching domain expiration for domain: {domain} from What\'s My DNS.')
                    return 'Error'
            print(f'Domain: {domain}, Expiry Date: {expiration_date}')
            return expiration_date
    return 'Invalid Domain'

# Path to your Excel file on OneDrive
excel_file = r'C:\Users\gauri.kothekar\OneDrive - Netsmartz LLC\DomainExp 14.xlsx'

# Load the existing workbook or create a new one if it doesn't exist
try:
    wb = load_workbook(excel_file)
except FileNotFoundError:
    wb = Workbook()

# Read the sheets into DataFrames
sheet_names = wb.sheetnames
df_dict = {}
for sheet_name in sheet_names:
    df_dict[sheet_name] = pd.read_excel(excel_file, sheet_name=sheet_name)

# Process each sheet
for sheet_name, df in df_dict.items():
    if 'Domain' not in df.columns:
        continue  # Skip sheets without the required column
    if 'Expiry' not in df.columns:
        df['Expiry'] = ''
    if 'Prod Type' not in df.columns:
        continue  # Skip sheets without the 'Prod Type' column
    
    # Update the expiry dates
    df['Expiry'] = df.apply(update_expiry, axis=1)
    
    # Get the worksheet
    ws = wb[sheet_name]

    # Remove existing data from the worksheet
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        ws.delete_rows(row[0].row)
    
    # Add header
    for col_num, column_title in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)
    
    # Add data rows
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Apply conditional formatting
    highlight_fill_red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    highlight_fill_orange = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    highlight_fill_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    highlight_fill_green = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

    today = datetime.now(timezone.utc)
    seven_days_later = today + timedelta(days=7)
    fifteen_days_later = today + timedelta(days=15)
    one_month_later = today + timedelta(days=30)

    expiry_col_idx = df.columns.get_loc('Expiry') + 1  # Excel columns are 1-based

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=expiry_col_idx)
        if cell.value and cell.value != 'Error':
            try:
                expiry_date = datetime.strptime(cell.value, '%Y-%m-%d')
                expiry_date = expiry_date.replace(tzinfo=timezone.utc)
                if today <= expiry_date <= seven_days_later:
                    cell.fill = highlight_fill_red
                elif today <= expiry_date <= fifteen_days_later:
                    cell.fill = highlight_fill_orange
                elif today <= expiry_date <= one_month_later:
                    cell.fill = highlight_fill_yellow
                else:
                    cell.fill = highlight_fill_green
            except ValueError:
                pass

# Remove the default sheet if it's empty (openpyxl creates a default sheet sometimes)
if 'Sheet' in wb.sheetnames and not wb['Sheet'].max_row:
    del wb['Sheet']

# Save the workbook
wb.save(excel_file)

print('Excel file updated with all sheets successfully.')
